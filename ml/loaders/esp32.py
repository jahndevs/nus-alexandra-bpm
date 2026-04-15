import numpy as np
from openpyxl import load_workbook

# ESP32 sampling rate, don't change
FS = 125


# load sessions from the ppg_dataset excel file
# returns list of dicts: {session_id, ppg, fs, metadata}
def load_sessions(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    sessions = {}
    for sid, ts, val, sbp, dbp, age, weight, height in ws.iter_rows(min_row=2, values_only=True):
        if sid is None:
            continue
        if sid not in sessions:
            sessions[sid] = {
                "samples": [],
                "metadata": {
                    "sbp": sbp, "dbp": dbp,
                    "age": age, "weight": weight, "height": height,
                },
            }
        sessions[sid]["samples"].append(val)

    result = []
    for sid, s in sessions.items():
        result.append({
            "session_id": sid,
            "ppg": np.array(s["samples"], dtype=float),
            "fs": FS,
            "metadata": s["metadata"],
        })
    return result
