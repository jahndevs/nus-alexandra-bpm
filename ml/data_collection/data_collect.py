"""
Data collection script that collects raw PPG sensor data from ESP32 for a specified duration, then asks for actual BP reading and some biometric data for storing in Excel.
"""

import serial
import serial.tools.list_ports
import time
import os
from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
EXCEL_FILE = os.path.join(DATA_DIR, "ppg_dataset.xlsx")
COLLECTION_DURATION = 30  # seconds
BAUD_RATE = 115200


def list_serial_ports():
    ports = serial.tools.list_ports.comports()
    if not ports:
        print("No serial ports found.")
        return None
    print("\nAvailable serial ports:")
    for i, port in enumerate(ports):
        print(f"  [{i}] {port.device} - {port.description}")
    while True:
        choice = input(f"\nSelect port [0-{len(ports)-1}]: ").strip()
        if choice.isdigit() and 0 <= int(choice) < len(ports):
            return ports[int(choice)].device
        print("Invalid selection.")


def collect_ppg(ser):
    print(f"\nCollecting PPG data for {COLLECTION_DURATION} seconds...")
    samples = []
    start = time.time()

    ser.reset_input_buffer()

    while time.time() - start < COLLECTION_DURATION:
        elapsed = time.time() - start
        remaining = COLLECTION_DURATION - elapsed

        line = ser.readline().decode("utf-8", errors="ignore").strip()
        if line:
            if len(samples) < 3:  # debug line
                print(f"  DEBUG raw line: {repr(line)}")
            if line.isdigit():
                timestamp_ms = int((time.time() - start) * 1000)
                samples.append((timestamp_ms, int(line)))

        # update progress
        if len(samples) % 100 == 0 and len(samples) > 0:
            print(f"  {remaining:.0f}s remaining | {len(samples)} samples", end="\r")

    print(f"\nDone! Collected {len(samples)} samples.")
    return samples


def get_metadata():
    print("\n--- Enter subject info ---")

    def ask_int(prompt):
        while True:
            val = input(prompt).strip()
            if val.isdigit():
                return int(val)
            print("  Please enter a whole number.")

    def ask_float(prompt):
        while True:
            val = input(prompt).strip()
            try:
                return float(val)
            except ValueError:
                print("  Please enter a number.")

    systolic = ask_int("Systolic BP (mmHg): ")
    diastolic = ask_int("Diastolic BP (mmHg): ")
    age = ask_int("Age (years): ")
    weight = ask_float("Weight (kg): ")
    height = ask_float("Height (cm): ")

    return {
        "systolic": systolic,
        "diastolic": diastolic,
        "age": age,
        "weight": weight,
        "height": height,
    }


def save_to_excel(samples, metadata):
    os.makedirs(DATA_DIR, exist_ok=True)

    headers = [
        "session_id", "timestamp_ms", "ppg_raw",
        "systolic", "diastolic", "age", "weight_kg", "height_cm",
    ]

    if os.path.isfile(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PPG Data"
        ws.append(headers)

    session_id = int(time.time())

    for ts, val in samples:
        ws.append([
            session_id, ts, val,
            metadata["systolic"], metadata["diastolic"],
            metadata["age"], metadata["weight"], metadata["height"],
        ])

    wb.save(EXCEL_FILE)
    print(f"\nSaved {len(samples)} rows to {EXCEL_FILE}")
    print(f"Session ID: {session_id}")


def main():
    print("=" * 40)
    print("  PPG Data Collection")
    print("=" * 40)

    port = list_serial_ports()
    if port is None:
        return

    print(f"\nOpening {port} at {BAUD_RATE} baud...")
    ser = serial.Serial(port, BAUD_RATE, timeout=1)
    time.sleep(2) # wait for ESP32 reset

    try:
        while True:
            input("\nPress ENTER to start collecting (or Ctrl+C to quit)...")
            samples = collect_ppg(ser)

            if not samples:
                print("ERROR: No data received.")
                continue

            metadata = get_metadata()
            save_to_excel(samples, metadata)

            again = input("\nCollect another session? [y/N]: ").strip().lower()
            if again != "y":
                break
    except KeyboardInterrupt:
        print("\nExiting.")
    finally:
        ser.close()
        print("Serial port closed.")


if __name__ == "__main__":
    main()
